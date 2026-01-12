from __future__ import annotations

from pathlib import Path
from time import time

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


def _next_free_path(filename: str) -> str:
    p = Path(filename)
    if not p.exists():
        return str(p)

    stem, suffix = p.stem, p.suffix
    parent = p.parent

    for i in range(1, 101):
        cand = parent / f"{stem}_{i}{suffix}"
        if not cand.exists():
            return str(cand)

    ts = int(time())
    return str(parent / f"{stem}_{ts}{suffix}")


def _safe_save_workbook(wb: Workbook, filename: str) -> str:
    try:
        wb.save(filename)
        return filename
    except PermissionError:
        alt = _next_free_path(filename)
        wb.save(alt)
        return alt


def _style_header(ws, headers: list[str]) -> None:
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"


def _auto_width(ws, headers: list[str], compact_cols: set[int] | None = None) -> None:
    compact_cols = compact_cols or set()

    for col_idx, h in enumerate(headers, start=1):
        max_len = len(str(h))
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=col_idx).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))

        limit = 22 if col_idx in compact_cols else 60
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, limit)


def _add_table(ws, name: str) -> None:
    ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    tbl = Table(displayName=name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tbl)


def write_excel_report(deposits_df: pd.DataFrame, sites_df: pd.DataFrame, filename: str) -> str:
    wb = Workbook()

    ws_deps = wb.active
    ws_deps.title = "USD Deposits"

    if deposits_df.empty:
        c = ws_deps["A1"]
        c.value = "No USD deposits found"
        c.font = Font(italic=True, color="808080")
    else:
        headers = list(deposits_df.columns)
        ws_deps.append(headers)
        for row in deposits_df.itertuples(index=False):
            ws_deps.append(list(row))

        _style_header(ws_deps, headers)

        if "AnnualRate" in headers:
            rate_col = headers.index("AnnualRate") + 1
            for r in range(2, ws_deps.max_row + 1):
                cell = ws_deps.cell(row=r, column=rate_col)
                cell.number_format = "0.0%"
                cell.alignment = Alignment(horizontal="right", vertical="center")

                v = cell.value
                if isinstance(v, (int, float)):
                    if v > 0.10:
                        cell.font = Font(color="E74C3C", bold=True)
                    elif v > 0.05:
                        cell.font = Font(color="F39C12", bold=True)

        if "SourceURL" in headers:
            url_col = headers.index("SourceURL") + 1
            for r in range(2, ws_deps.max_row + 1):
                cell = ws_deps.cell(row=r, column=url_col)
                url = cell.value
                if isinstance(url, str) and url.startswith(("http://", "https://")):
                    cell.hyperlink = url
                    cell.font = Font(color="0563C1", underline="single")
                    cell.value = "Open"

        _add_table(ws_deps, "DepositsUSD")
        _auto_width(ws_deps, headers)

    ws_sites = wb.create_sheet("Site Status", 0 if deposits_df.empty else 1)

    if sites_df.empty:
        c = ws_sites["A1"]
        c.value = "No sites processed"
        c.font = Font(italic=True, color="808080")
    else:
        headers = list(sites_df.columns)
        ws_sites.append(headers)
        for row in sites_df.itertuples(index=False):
            ws_sites.append(list(row))

        _style_header(ws_sites, headers)

        if "Result" in headers:
            result_col = headers.index("Result") + 1
            for r in range(2, ws_sites.max_row + 1):
                cell = ws_sites.cell(row=r, column=result_col)
                v = str(cell.value or "")
                if "OK" in v:
                    cell.font = Font(color="27AE60", bold=True)
                elif ("ERROR" in v) or ("BLOCKED" in v):
                    cell.font = Font(color="E74C3C", bold=True)
                elif ("NO_" in v) or ("JS_" in v):
                    cell.font = Font(color="F39C12")

        _add_table(ws_sites, "SiteStatus")
        compact = set()
        for name in ("HTTPStatus", "RowsFound"):
            if name in headers:
                compact.add(headers.index(name) + 1)
        _auto_width(ws_sites, headers, compact_cols=compact)

    return _safe_save_workbook(wb, filename)


def write_csv(df: pd.DataFrame, filename: str) -> str:
    try:
        df.to_csv(filename, index=False, encoding="utf-8-sig")
        return filename
    except PermissionError:
        alt = _next_free_path(filename)
        df.to_csv(alt, index=False, encoding="utf-8-sig")
        return alt
